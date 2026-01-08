import pandas as pd
from bs4 import BeautifulSoup
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# --------------------- CONFIGURATION ---------------------
YOUR_EMAIL = os.environ.get('EMAIL_USER')
APP_PASSWORD = os.environ.get('EMAIL_PASS')
TO_EMAIL = "priyansh.khandelwal@rosierfoods.com"

EXCEL_FILE = "blinkit_rosier_products.xlsx"
HTML_FILE = "blinkit.html"

# Security Check
if not YOUR_EMAIL or not APP_PASSWORD:
    pass 

# --------------------- SCRAPING LOGIC ---------------------
print("Reading HTML file...")
try:
    with open(HTML_FILE, 'r', encoding='utf-8') as file:
        html_content = file.read()
except FileNotFoundError:
    print(f"Error: '{HTML_FILE}' file nahi mili!")
    exit()

soup = BeautifulSoup(html_content, 'html.parser')

product_containers = soup.find_all('div', attrs={'role': 'button', 'tabindex': '0'})
print(f"Total containers found: {len(product_containers)}")

product_details = []

for container in product_containers:
    # 1. Title Extraction
    title_div = container.find('div', class_=re.compile(r'tw-text-300.*tw-font-semibold.*tw-line-clamp-2'))
    if not title_div:
        continue
    
    product_name = title_div.get_text(strip=True)
    
    if 'rosier' not in product_name.lower():
        continue

    # ============================================================
    # 2. LINK EXTRACTION (ULTIMATE FALLBACK LOGIC)
    # ============================================================
    product_url = None
    link_tag = None
    
    # LEVEL 1: Container ke andar dhundo
    link_tag = container.find('a', href=True)
    
    # LEVEL 2: Container ke parent me dhundo
    if not link_tag:
        link_tag = container.find_parent('a', href=True)

    # LEVEL 3 (Brahmastra): Pure Title Text se Link dhundo
    # Logic: Pure page me aisa 'a' tag dhundo jiske andar ye Product Name likha ho
    if not link_tag:
        # Hum us element ko dhundenge jisme exact product name hai
        text_element = soup.find(string=re.compile(re.escape(product_name)))
        if text_element:
            # Us text ka parent 'a' tag dhundo
            link_tag = text_element.find_parent('a', href=True)

    # URL Create karna
    if link_tag:
        href_val = link_tag['href'].strip()
        if href_val.startswith("http"):
            product_url = href_val
        elif href_val.startswith("/"):
            product_url = "https://blinkit.com" + href_val
        else:
            product_url = "https://blinkit.com/" + href_val
            
    # ============================================================

    # 3. Variant
    variant = "-"
    next_div = title_div.find_next_sibling('div')
    if next_div:
        text = next_div.get_text(strip=True)
        if any(unit in text.lower() for unit in ['kg', 'g', 'ml', 'l', 'pack', 'piece', 'pcs']):
            variant = text
    if variant == "-":
        container_text = container.get_text(separator=" ", strip=True)
        match = re.search(r'(\d+(?:\.\d+)?\s*(?:kg|g|ml|l|piece|pack|pcs|bottle|jar|box|kgx|gx|ltr|litre|kilogram))\b', container_text, re.IGNORECASE)
        if match:
            variant = match.group(1).strip()
    if variant == "-" and any(unit in product_name.lower() for unit in ['kg', 'g', 'ml', 'l']):
        words = product_name.split()
        for i in range(len(words)-1, -1, -1):
            if any(unit in words[i].lower() for unit in ['kg', 'g', 'ml', 'l', 'pack', 'piece']):
                start = max(0, i-1)
                variant = ' '.join(words[start:i+1])
                break

    # 4. Stock & Price
    text_content = container.get_text(separator=" ", strip=True).lower()
    has_out_stock = any(phrase in text_content for phrase in ['out of stock', 'outofstock', 'currently unavailable'])
    stock_status = "Out of Stock" if has_out_stock else "In Stock"

    price = "-"
    price_div = container.find('div', class_=re.compile(r'tw-text-200.*tw-font-semibold'))
    if price_div:
        price = price_div.get_text(strip=True)
    else:
        possible_price = container.find(string=re.compile(r'^₹[0-9,]+$'))
        if possible_price:
            price = possible_price.strip()

    product_details.append({
        "Title": product_name,
        "Variant": variant.strip(),
        "Price": price,
        "Stock": stock_status,
        "Hidden_URL": product_url
    })

# --------------------- EXCEL GENERATION ---------------------
if product_details:
    df = pd.DataFrame(product_details)
    df.to_excel(EXCEL_FILE, index=False)
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        title_cell = row[0] # Column A
        url_cell = row[4]   # Column E
        
        if url_cell.value:
            title_cell.hyperlink = url_cell.value
            title_cell.font = Font(color="0000FF", underline="single")
        else:
            # Agar abhi bhi Red hai, to file me link hi nahi hai
            title_cell.font = Font(color="FF0000", bold=True)
            print(f"Warning: No Link found for {title_cell.value}")

    ws.delete_cols(5)
    wb.save(EXCEL_FILE)
    print(f"Data saved to → {EXCEL_FILE}")
else:
    print("No Rosier products found.")
    exit()

# --------------------- EMAIL SENDING ---------------------
if not YOUR_EMAIL or not APP_PASSWORD:
    print("Skipping Email (No Credentials).")
    exit()

print("Sending Email...")
msg = MIMEMultipart()
msg['From'] = YOUR_EMAIL
msg['To'] = TO_EMAIL
msg['Subject'] = 'Blinkit - Latest Rosier Products'
body = f"Hi Automailer PFA Rosier blinkit products.\nTotal: {len(product_details)}"
msg.attach(MIMEText(body, 'plain'))
with open(EXCEL_FILE, 'rb') as attachment:
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', f'attachment; filename={EXCEL_FILE}')
msg.attach(part)
try:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(YOUR_EMAIL, APP_PASSWORD)
    server.send_message(msg)
    server.quit()
    print("Email Sent!")
except Exception as e:
    print(f"Email Failed: {e}")
